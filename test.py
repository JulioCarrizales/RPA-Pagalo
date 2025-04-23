import pyautogui
import time
import os
import webbrowser
import pyperclip
import csv
import openpyxl
from whatsapp import whatsapp

pyautogui.FAILSAFE = True

COORDENADAS = {
    "usuario":  (811, 287),
    "contraseña": (750,322),
    "botón_de_inicio_sesión": (796,355),
    "Primer_filtro": (531,170),
    "OperacionesRechazadas": (531,206),
    "NoProcesadas":(725,280),
    "Bug": (723,207),
    "Fecha_1": (322,319),
    "Boton_Consultar_datos":(1273,336),
    "NrTicker_1":(226,495),
    "Reload_Fila_1":(1429,494),
    "NrTicker_2":(227,519),
    "Reload_Fila_2":(1430,518),
    "NrTicker_3":(221,543),
    "Reload_Fila_3":(1431,543),
    "NrTicker_4":(224,568),
    "Reload_Fila_4":(1429,568),
    "NrTicker_5":(221,593),
    "Reload_Fila_5":(1432,595),
    "NrTicker_6":(223,619),
    "Reload_Fila_6":(1429,622),
    "NrTicker_7":(224,644),
    "Reload_Fila_7":(1431,645),
    "NrTicker_8":(222,669),
    "Reload_Fila_8":(1431,669),
    "NrTicker_9":(222,694),
    "Reload_Fila_9":(1430,694),
    "NrTicker_10":(222,719),
    "Reload_Fila_10":(1431,721),
    "Elimiar_Fila_1":(1388,494),
    "Elimiar_Fila_2":(1389,519),
    "Elimiar_Fila_3":(1388,545),
    "Elimiar_Fila_4":(1388,548),
    "Elimiar_Fila_5":(1388,594),
    "Elimiar_Fila_6":(1388,621),
    "Elimiar_Fila_7":(13889,646),
    "Elimiar_Fila_8":(1389,670),
    "Elimiar_Fila_9":(1389,693),
    "Elimiar_Fila_10":(1388,719),
    "Confirmar_eliminacion":(912,508),
    "Close_window":(1571,19),
    "NextPage": (845,743),
}

# CREDENCIALES
usuario = "pmce"
contrasena = "abril17." 

# FUNCIONES PARA MOVIMIENTO DE RPA

def mover_mouse(x, y, delay=0.2):
    """Mueve el mouse a una posición y hace clic."""
    pyautogui.moveTo(x, y, duration=delay)

def mover_mouse_y_clic(x, y, delay=0.2):
    """Mueve el mouse a una posición y hace clic."""
    pyautogui.moveTo(x, y, duration=delay)
    pyautogui.click()

def doubleclick(x, y, delay=0.2):
    """Mueve el mouse a una posición y hace clic."""
    pyautogui.moveTo(x, y, duration=delay)
    pyautogui.doubleClick()

def escribir_texto(texto, delay=0.1):
    """Escribe texto con un retraso entre caracteres."""
    pyautogui.write(texto, interval=delay)

def copiar_y_restar():
    """Copia un número al portapapeles, le resta 5 y devuelve el resultado."""
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(1)  # Esperar un momento para asegurarse de que el texto se copie
    numero_copiado = pyperclip.paste()
    try:
        numero = int(numero_copiado)
        resultado = numero - 5
        escribir_texto(str(resultado))
        return resultado
    except ValueError:
        print("El contenido copiado no es un número válido.")
        return None

def copiar_y_guardar_en_excel(reiniciar_contador=False):
    """Copia un número al portapapeles y lo guarda en un archivo Excel y devuelve el contador."""
    pyautogui.hotkey('ctrl', 'c')
    time.sleep(1)  # Esperar un momento para asegurarse de que el texto se copie
    numero_copiado = pyperclip.paste()
    print(f"Contenido copiado: {numero_copiado}")  # Agregar impresión de depuración
    try:
        numero = int(numero_copiado)
        ruta_excel = "C:\\Users\\Administrador\\Desktop\\RPA-Pagalo\\Tickets.xlsx"
        if not os.path.exists(ruta_excel):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Hoja1"
            ws.append(["Nr_Ticket", "Contador"])
        else:
            wb = openpyxl.load_workbook(ruta_excel)
            ws = wb["Hoja1"]
        
        contador = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == numero:
                contador = row[1] + 1
                if reiniciar_contador:
                    contador = 0
                break
        
        ws.append([numero, contador])
        wb.save(ruta_excel)
        print(f"NrTicket: {numero}, Contador actualizado: {contador}")  # Agregar impresión de validación
        return numero, contador
    except ValueError:
        print("El contenido copiado no es un número válido.")
        return None, None

def escribir_log(mensaje):
    """Escribe un mensaje en el archivo de log."""
    ruta_log = "C:\\Users\\Administrador\\Desktop\\RPA-Pagalo\\log.txt"

    with open(ruta_log, "a") as log_file:
        log_file.write(f"{time.strftime('%Y-%m-%d %H:%M:%S')} - {mensaje}\n")

# FUNCIONES PARA LAS ACCIONES DEL RPA

def abrir_pagina():
    """Abre el enlace en el navegador predeterminado."""
    webbrowser.open("http://10.7.10.169:9080/MonitorMultipagos/usuarios/sessionTerminada.action")
    time.sleep(5)
    pyautogui.click(x=360, y=883)  # Hacer clic en la página para asegurarse de que esté activa
    time.sleep(2)
    pyautogui.click(x=410, y=884)  # Hacer clic en la página para asegurarse de que esté activa
    time.sleep(2)
    pyautogui.hotkey('win','up')
    time.sleep(3)
    mover_mouse_y_clic(*COORDENADAS["usuario"])
    time.sleep(2)
    escribir_texto(usuario)
    time.sleep(2)
    mover_mouse_y_clic(*COORDENADAS["contraseña"])
    time.sleep(2)
    escribir_texto(contrasena)
    time.sleep(2)
    mover_mouse_y_clic(*COORDENADAS["botón_de_inicio_sesión"])
    time.sleep(5)

def Operaciones_NoProcesadas():
    mover_mouse_y_clic(*COORDENADAS["Primer_filtro"])
    time.sleep(3)
    mover_mouse(*COORDENADAS["OperacionesRechazadas"])
    time.sleep(2)
    mover_mouse(*COORDENADAS["Bug"])
    time.sleep(2)
    mover_mouse_y_clic(*COORDENADAS["NoProcesadas"])
    time.sleep(2)
    doubleclick(*COORDENADAS["Fecha_1"])
    time.sleep(2)
    mover_mouse_y_clic(*COORDENADAS["Boton_Consultar_datos"])
    time.sleep(20)

def fila_1():
    doubleclick(*COORDENADAS["NrTicker_1"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_1"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador=True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_1"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_2():
    doubleclick(*COORDENADAS["NrTicker_2"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_2"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador=True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_2"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_3():
    doubleclick(*COORDENADAS["NrTicker_3"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_3"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador=True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_3"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_4():
    doubleclick(*COORDENADAS["NrTicker_4"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_4"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador=True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_4"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_5():
    doubleclick(*COORDENADAS["NrTicker_5"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_5"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador=True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_5"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_6():
    doubleclick(*COORDENADAS["NrTicker_6"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_6"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador = True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_6"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_7():
    doubleclick(*COORDENADAS["NrTicker_7"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_7"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador = True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_7"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_8():
    doubleclick(*COORDENADAS["NrTicker_8"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_8"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador = True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_8"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_9():
    doubleclick(*COORDENADAS["NrTicker_9"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_9"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador = True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_9"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def fila_10():
    doubleclick(*COORDENADAS["NrTicker_10"])
    time.sleep(2)
    numero, contador = copiar_y_guardar_en_excel()
    time.sleep(2)
    if contador >= 2:
        mover_mouse_y_clic(*COORDENADAS["Elimiar_Fila_10"])
        escribir_log(f"NrTicket {numero} eliminado.")
        copiar_y_guardar_en_excel(reiniciar_contador = True)
        mover_mouse_y_clic(*COORDENADAS["Confirmar_eliminacion"])
    else:
        mover_mouse_y_clic(*COORDENADAS["Reload_Fila_10"])
        escribir_log(f"NrTicket {numero} recargado.")
    time.sleep(10)

def cambiar_red(red):
    """Cambia la red de internet a la especificada."""
    print(f"Cambiando a la red: {red}")
    os.system(f'netsh wlan connect name="{red}"')
    time.sleep(10)  # Esperar a que la conexión se establezca

def crear_excel_tickets():
    """Crea el archivo Tickets.xlsx si no existe."""
    ruta_excel = "C:\\Users\\Administrador\\Desktop\\RPA-Pagalo\\Tickets.xlsx"
    if not os.path.exists(ruta_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        ws.append(["Nr_Ticket", "Contador"])
        wb.save(ruta_excel)
        print(f"Archivo creado: {ruta_excel}")

def eliminar_excel_tickets():
    """Elimina el archivo Tickets.xlsx si existe."""
    ruta_excel = "C:\\Users\\Administrador\\Desktop\\RPA-Pagalo\\Tickets.xlsx"
    if os.path.exists(ruta_excel):
        os.remove(ruta_excel)
        print(f"Archivo eliminado: {ruta_excel}")

def main():
    while True:  # Bucle infinito para reiniciar el proceso en caso de error
        try:
            crear_excel_tickets()  # Crear el archivo al inicio
            cambiar_red("SBDIR")  # Cambiar a la red SBDIR para usar WhatsApp
            whatsapp()
            cambiar_red("BNCORP")  # Cambiar a la red BNCORP para el resto del proceso
            abrir_pagina()
            Operaciones_NoProcesadas()
            start_time = time.time()  # Registrar el tiempo de inicio
            while True:
                fila_1()
                fila_2()
                fila_3()
                fila_4()
                fila_5()
                fila_6()
                fila_7()
                fila_8()
                fila_9()
                fila_10()
                time.sleep(5)
                mover_mouse_y_clic(*COORDENADAS["NextPage"])
                time.sleep(15)

                # Verificar si han pasado 3 horas
                if time.time() - start_time >= 3 * 60 * 60:  # 3 horas en segundos
                    mover_mouse_y_clic(*COORDENADAS["Close_window"])
                    time.sleep(5)
                    eliminar_excel_tickets()  # Eliminar el archivo al final del proceso
                    time.sleep(5)
                    break  # Salir del bucle interno para reiniciar el proceso
        except Exception as e:
            print(f"Error detectado: {e}. Reiniciando el proceso...")
            time.sleep(10)  # Esperar antes de reiniciar
            continue  # Reiniciar el proceso desde el principio

if __name__ == "__main__":
    main()
